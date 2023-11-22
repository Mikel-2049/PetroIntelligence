from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

################################################################################################################################
################################################################################################################################
################################################################################################################################
################################################################################################################################

daily_file = 'Petrointelligance-Data_8_30.png'

################################################################################################################################
################################################################################################################################
################################################################################################################################
################################################################################################################################


# Set the desired width and height for the screenshot
desired_width = 1440
desired_height = 9824

# Create Firefox options with the local file detector capability
options = Options()
options.set_capability("acceptInsecureCerts", True)  # This might be needed for local files with HTTPS

# Create the browser instance with the desired options
browser = webdriver.Firefox(options=options)

# Load the local file using the file:// protocol
local_file_path = "file:///C:/Users/Mr.Harbor/OneDrive%20-%20HARBOR%20INTELLIGENCE%20SC/Mikel%20González/Petro_Dev/Gas_viewer.html"
browser.get(local_file_path)

# Wait for the content to load using an explicit wait
wait = WebDriverWait(browser, 60)  # Wait for up to 60 seconds
element = wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))

# Get the total height of the page
total_height = browser.execute_script("return document.body.scrollHeight")

# Set the browser window size to match the desired dimensions
browser.set_window_size(desired_width, total_height)

# Give additional time for the page to fully load (adjust as needed)
time.sleep(5)  # Wait for an additional 5 seconds

# Take a screenshot of the entire page
################################################################################################################################
################################################################################################################################
################################################################################################################################
################################################################################################################################
screenshot_path = daily_file
################################################################################################################################
################################################################################################################################
################################################################################################################################
################################################################################################################################
browser.save_screenshot(screenshot_path)

# Close the browser
browser.quit()




import cv2
import pytesseract
import re
import numpy as np
from itertools import zip_longest
from openpyxl import Workbook, load_workbook
from datetime import date

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract'

def hex_to_bgr(hex_color):
    # Convert hex to RGB
    rgb = [int(hex_color[i:i+2], 16) for i in (1, 3, 5)]
    # Convert RGB to BGR
    return tuple(reversed(rgb))
################################################################################################################################
################################################################################################################################
################################################################################################################################
################################################################################################################################
################################################################################################################################
image_path = daily_file
image = cv2.imread(image_path)
################################################################################################################################
################################################################################################################################
################################################################################################################################
################################################################################################################################
################################################################################################################################
# Define the exact BGR colors
colors = {
    'green': hex_to_bgr('#489871'),
    'red': hex_to_bgr('#D53A58'),
    'black': hex_to_bgr('#5B5B5B')
}

# Dictionary to hold the extracted prices grouped by color
prices_by_color = {'green': [], 'red': [], 'black': []}

for color_name, color_value in colors.items():
    # Create a mask for the exact color
    lower_bound = np.array(color_value, dtype=np.uint8)
    upper_bound = np.array(color_value, dtype=np.uint8)
    mask = cv2.inRange(image, lower_bound, upper_bound)
    
    # Find contours in the masked image
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    for contour in contours:
        # Get the bounding box of the contour
        x, y, w, h = cv2.boundingRect(contour)
        
        # Crop the region containing the price
        region = image[y:y+h, x:x+w]
        
        # Apply OCR to the region
        text = pytesseract.image_to_string(region, config='--oem 3 --psm 6')
        
        # Match the pattern of 2 digits before and 2 digits after a decimal point
        match = re.search(r'\d{2}\.\d{2}', text)
        if match:
            price = match.group()
            prices_by_color[color_name].append((x, y, price))

# Sort the prices by y-coordinate (assumed to represent the order of states)
sorted_green = sorted(prices_by_color['green'], key=lambda p: p[1])
sorted_red = sorted(prices_by_color['red'], key=lambda p: p[1])
sorted_black = sorted(prices_by_color['black'], key=lambda p: p[1])

# Load the existing workbook
workbook_path = 'numbers.xlsx'
workbook = load_workbook(workbook_path)
worksheet = workbook['Sheet']

# Find the last row with data in column B
last_row = worksheet.max_row
for row in reversed(range(1, last_row + 1)):
    if worksheet.cell(row=row, column=2).value is not None:
        last_row = row + 1
        break

# Counter for the total number of prices
total_count = 0

# Write the extracted prices starting from column B in the last empty row
current_column = 2  # Column B
for green, red, black in zip_longest(sorted_green, sorted_red, sorted_black, fillvalue=(None, None, None)):
    for price in (green[2], red[2], black[2]):
        if price:
            # Write the price to the current column in the last empty row
            worksheet.cell(row=last_row, column=current_column, value=price)
            total_count += 1
            current_column += 1

# Save the workbook
workbook.save(workbook_path)

print(f"Total Count: {total_count}")